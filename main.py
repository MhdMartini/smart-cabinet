import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import time
from datetime import datetime
from rpi_camera import CabinetCamera


class Shoebox:
    # Shoebox class
    columns = ["Time", "Action", "User", "URL"]

    def __init__(self, tag_num, box_num):
        self.tag_number = tag_num  # Tag number
        self.number = box_num  # Shoebox number

    def open_ticket(self, borrower):
        # Write a new line to the excel sheet
        borrow_time = datetime.now().strftime("%m/%d/%Y-%H:%M:%S")

        # Account for the case if the excel file or the sheet do not exist
        try:
            # If the file does not exist, create new one
            log = load_workbook("log.xlsx")
            try:
                # If the excel file exists but the sheet does not
                shoebox = log[self.number]
            except KeyError:
                shoebox = log.create_sheet(self.shoebox_number)
                shoebox["A1"], shoebox["B1"], shoebox["C1"], shoebox["D1"] = self.columns  # Account f
        except FileNotFoundError:
            log = Workbook()
            log["Sheet"].title = self.number
            shoebox = log[self.number]
            shoebox["A1"], shoebox["B1"], shoebox["C1"], shoebox["D1"] = self.columns

        # Add open ticket info
        # TODO: Get actual values of "borrow_time" and "borrower" from RFID reads
        row = [borrow_time, "Borrow", borrower, "url"]
        last_row = shoebox.max_row
        for idx, value in enumerate(row, start=1):
            # Output the borrow status. Skip a line after the previous return
            shoebox.cell(row=last_row + 2, column=idx, value=value)

        # Save file
        log.save("log.xlsx")

    def close_ticket(self, returner):
        return_time = datetime.now().strftime("%m/%d/%Y-%H:%M:%S")

        # Load current sheet
        log = load_workbook("log.xlsx")
        # HINT: May need to string typecast
        shoebox = log[self.number]

        # Add closed ticket info
        # TODO: Get actual values of "return_time" and "returner" from RFID reads
        row = [return_time, "Return", returner, "url"]
        last_row = shoebox.max_row
        for idx, value in enumerate(row, start=1):
            shoebox.cell(row=last_row + 1, column=idx, value=value)

        # Save file
        log.save("log.xlsx")

    def __repr__(self):
        return f"Shoebox {self.number}"


class SmartCabinet:
    Admin = False

    # TODO: ADD:
    # camera = CabinetCamera()
    RFID_tags = {}  # Contains Shoebox objects indexed by RFID tags
    permission_list = {}  # Contains allowed students' names indexed by RFID numbers
    admins = {}  # Contains admins' names indexed by RFID numbers

    # Two inventory lists to be compared after each transaction
    inventory = {}
    current_inventory = {}

    last_user = (None, None)  # (ID, name)

    # TODO: assign the value of "door" based on an IO Interrupt
    door = False  # True: Open, False: Closed

    def __init__(self):
        # Sync permission_list and RFID_tags with the online sheets
        self.update_sheets()

        # TODO Block
        # Get an initial list of all tag numbers (maybe read them in the cabinet) Then, go in the main loop
        # where you wait for a valid id to be scanned and door to be opened
        # TODO: Interrupt format:
        #  "GPIO.add_event_detect(Pin#, GPIO.FALLING, callback=my_callback, bouncetime=ms)"
        # TODO: Consider Admin vs. Regular mode:
        #  * Admin mode: the Admin scans his/her ID for 2 seconds, and then they leave the door open for
        #  students in the lab to grab their boxes after only scanning their IDs
        #  * Regular mode: a user has to scan their ID which unlocks the door,
        #  they can then take box(es) and close the door back.
        #  Admin scanning in normal mode (one scan) may not trigger output (ASK Recepient)

        self.initial_scan()  # Obtain inventory list, fill
        #
        # GPIO.add_event_detect(Pin#, GPIO.FALLING, callback=self.id_scanned, bouncetime=ms)

        # Get into ID scan loop
        self.main_loop()

    def update_sheets(self):
        students_sheet_id = "1DGLYKbAhM8OrI-BGUoaepaMbE2_euwzOpLXJFFsUMyQ"
        tags_sheet_id = "1VohFgEAXe1nWDXxFwre061KpLg7vxDcj989imjGoaIY"
        admin_sheet_id = "1b0jq554AXPzvwx5dbxkpcfk8Ne5HsQBM4LWhLFu5Dww"

        # Import students online spreadsheet
        # Create permission_list dictionary:
        # {RFID num 1: Student 1 name,
        #  RFID num 2: Student 2 name,
        #  etc.}
        # https://docs.google.com/spreadsheets/d/1DGLYKbAhM8OrI-BGUoaepaMbE2_euwzOpLXJFFsUMyQ/edit?usp=sharing
        students = pd.read_csv(f"https://docs.google.com/spreadsheets/d/{students_sheet_id}/export?format=csv")
        self.permission_list = {RFID_num: name for RFID_num, name in zip(
            students["RFID Number"], students["Student Name"]
        )}

        # Import RFID tags (shoeboxes) from the online spreadsheet
        # Create an RFID_tags dictionary:
        # {tag1 : Shoebox object 1,
        # {tag2 : Shoebox object 2,
        # etc.}
        # https://docs.google.com/spreadsheets/d/1VohFgEAXe1nWDXxFwre061KpLg7vxDcj989imjGoaIY/edit?usp=sharing
        tags = pd.read_csv(f"https://docs.google.com/spreadsheets/d/{tags_sheet_id}/export?format=csv")
        self.RFID_tags = {tag_num: Shoebox(tag_num, str(box_num)) for tag_num, box_num in zip(
            tags["RFID Tag"], tags["Shoebox Number"]
        )}

        # Import admins online spreadsheet
        # Create admins dictionary:
        # {RFID num 1: Admin 1 name,
        #  RFID num 2: Admin 2 name,
        #  etc.}
        # https://docs.google.com/spreadsheets/d/1b0jq554AXPzvwx5dbxkpcfk8Ne5HsQBM4LWhLFu5Dww/edit?usp=sharing
        # Anyone who has the link is an editor to the spreadsheet
        admins = pd.read_csv(f"https://docs.google.com/spreadsheets/d/{admin_sheet_id}/export?format=csv")
        self.admins = {RFID_num: name for RFID_num, name in zip(
            admins["RFID Number"], admins["Admin Name"]
        )}

    def initial_scan(self):
        # TODO: Scan the RFID tags in the Cabinet and return a list of tag numbers
        self.inventory = {"Tag1", "Tag2", "Tag3", "Tag4", "etc"}

    def inventory_scan(self):
        # TODO: Add method to read serial data from RFID reader and store the scanned
        #  tags (shoeboxes) in self.current_inventory
        self.current_inventory = {"Tag1", "Tag2", "Tag3", "etc"}

    def ID_scanned(self):
        RFID_number = 0xFF  # Some number

    def main_loop(self):
        while True:
            # Reset Admin status
            self.Admin = False

            # TODO: Consider interrupt instead of polling
            # TODO: Scan for ID
            id = "something"
            # Keep scanning until an ID is scanned
            # If an ID is scanned, validate the ID number, id
            if not id:
                time.sleep(0.1)
                continue
            # TODO: Consider outputting failed attempts to open the cabinet
            student = self.permission_list.get(id)
            admin = self.admins.get(id)
            if not (student or admin):
                # If ID is not valid
                # TODO: Do something
                time.sleep(0.3)
                continue

            # Only get here when it's either a student or an admin
            # TODO:
            #  Beep
            if admin:
                beeps = 1
                for _ in range(2):
                    time.sleep(1)
                    # TODO: Scan for ID
                    # If ID is still there, repeat, if it's still there, sync sheets
                    # else trigger Admin mode
                    if id:
                        # TODO: Beep again
                        beeps += 1
                        self.Admin = not self.Admin
                    else:
                        break
                if beeps == 3:
                    self.update_sheets()
                    continue
            # Only get here when valid ID is scanned and self.Admin is properly set
            # TODO: Open the lock, monitor if the door will be open for "t" seconds
            #  if the door was not open, go back to main loop.
            #  Else, open camera and wait for the door to be closed. When door
            #  is closed, scan inventory; compare inventory to identify the missing/added boxes,
            #  then open/close tickets and save the video accordingly
            self.last_user = (id, student or admin)
            self.lock(False)
            t = time()

            # Wait until the door is open or 5 seconds have passed
            # If the door was open, start camera and wait for the door to be closed
            # If timeout, go back to main loop
            while (not self.door) or (time() < t + 5):
                continue
            if not self.door:
                continue

            # Only get here if the door was open
            # Wait until the door is closed or a timeout period occurs
            # After the time out period, continue as if the door was closed, and notify
            # an Admin (Consult client)
            # While waiting for the door to be closed, start video recording
            #
            pass  # Marker
            # TODO: ADD:
            self.camera.start()
            t = time()
            while self.door and (time() < t + 20):
                continue

            # Only get here when the door was opened then closed
            # TODO: Lock the door, stop camera and save the video.
            #  Videos will be saved in a folder, which saves up to seven days activity
            #  E.g. a video associated with shoebox number 25 will be saved under C:\{PATH}\videos\25
            # Scan inventory (modifies the self.current_inventory attribute)
            # Get the different items between the two sets by performing an XOR between them
            # If there are differences, open/close tickets according to where a tag came from
            # else go back to main loop
            # Finally, update the inventory list according to the new changes
            # TODO: ADD:
            self.camera.stop()
            self.lock()
            self.inventory_scan()
            different_tags = self.inventory ^ self.current_inventory
            if not different_tags:
                # If valid ID was scanned, the door opened and closed, but no item was borrowed or returned,
                continue
            # If items were borrowed/received
            self.handle_tickets(different_tags)
            self.inventory = self.current_inventory.copy()

            # go back to main loop
            continue

    def lock(self, lock_=True):
        # lock(True) locks the door; lock(False) unlocks the door
        # Lock or Unlock the door
        if lock_:
            # Lock the door
            return
        # Unlock the door

    def handle_tickets(self, tags):
        # Loop through the different tags and either open or close the ticket according to where
        # a tag came from.

        for tag in tags:
            box = self.RFID_tags[tag]
            user = self.last_user[1]

            if tag in self.inventory:
                # Item was borrowed
                # TODO: Consider creating a user class
                box.open_ticket(user)
            else:
                # Item was returned
                box.close_ticket(user)
        # TODO: Consider handling newly added tags (adding them to the system)
        #  Or, reconsider the system by which the list of all tags is obtained the first time
        #  Consider dynamic adding instead of having to fill an Excel sheet and having to restart the system
