from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime


class Shoebox:
    # Shoebox class
    columns = ["Time", "Action", "User", "URL"]  # Column names in log excel file
    idx = 0             # The index of the excel file to be filled, incase one file fills up
    MAX_ROWS = 5000     # Max rows in an excel sheet before we consider it filled up

    def __init__(self, tag_num, box_num):
        self.tag_number = tag_num  # Tag number
        self.number = box_num  # Shoebox number

    def open_ticket(self, borrower):
        # Write a new line to the excel sheet
        borrow_time = datetime.now().strftime("%m/%d/%Y-%H:%M:%S")

        # Account for the case if the excel file or the sheet do not exist
        try:
            # If the file does not exist, create new one
            log = load_workbook(f"log{self.idx}.xlsx")
            try:
                # If the excel file exists but the sheet does not, create the new sheet named after the shoebox
                # If the sheet exists, but the max number of rows was reached, start a new excel file from scratch
                shoebox = log[self.number]
                last_row_check = shoebox.max_row > self.MAX_ROWS
                if last_row_check:
                    # Act as if the whole file was not found, i.e., create a new file and sheet
                    self.idx += 1
                    raise FileNotFoundError

            except KeyError:
                shoebox = log.create_sheet(self.number)
                shoebox["A1"], shoebox["B1"], shoebox["C1"], shoebox["D1"] = self.columns
        except FileNotFoundError:
            log = Workbook()
            log["Sheet"].title = self.number
            shoebox = log[self.number]
            shoebox["A1"], shoebox["B1"], shoebox["C1"], shoebox["D1"] = self.columns
        finally:
            last_row = shoebox.max_row

        # Add open ticket info
        # TODO: Get actual values of "borrow_time" and "borrower" from RFID reads
        row = [borrow_time, "Borrow", borrower, "url"]
        for idx, value in enumerate(row, start=1):
            # Output the borrow status. Skip a line after the previous return
            shoebox.cell(row=last_row + 2, column=idx, value=value)

        # Save file
        log.save(f"log{self.idx}.xlsx")

    def close_ticket(self, returner):
        return_time = datetime.now().strftime("%m/%d/%Y-%H:%M:%S")

        # Load current sheet
        log = load_workbook("log.xlsx")
        # HINT: May need to string typecast
        shoebox = log[self.number]

        # Add closed ticket info
        # TODO: Get actual values of "return_time" and "returner" from RFID reads
        row = [return_time, "Return", returner, "url"]
        last_row = shoebox.max_row
        for idx, value in enumerate(row, start=1):
            shoebox.cell(row=last_row + 1, column=idx, value=value)

        # Save file
        log.save("log.xlsx")

    def __repr__(self):
        return f"Shoebox {self.number}"

